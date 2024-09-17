from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
import hashlib  # for password hashing

app = Flask(__name__)
CORS(app)

# Replace with your spreadsheet file path
spreadsheet_path = 'D:/code/Flutter/GO/DATA/users.xlsx'

# Helper function to hash passwords


def hash_password(password):
    """Hashes a password using SHA-256."""
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

# Authentication Endpoint


@app.route('/login', methods=['POST'])
def login():
    email = request.json.get('email')
    password = request.json.get('password')

    if not email or not password:
        return jsonify({'success': False, 'message': 'Email and password are required'}), 400

    # Load workbook
    try:
        workbook = openpyxl.load_workbook(spreadsheet_path)
    except FileNotFoundError:
        return jsonify({'success': False, 'message': 'User database not found'}), 500

    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2):
        if row[0].value == email:
            hashed_password = row[1].value
            if hashed_password == hash_password(password):
                # Send the username
                return jsonify({'success': True, 'user_name': row[2].value})
            else:
                return jsonify({'success': False, 'message': 'Incorrect password'}), 401

    return jsonify({'success': False, 'message': 'User not found'}), 404

# Sign Up Endpoint


@app.route('/signup', methods=['POST'])
def signup():
    email = request.json.get('email')
    password = request.json.get('password')
    user_name = request.json.get('user_name')

    if not email or not password or not user_name:
        return jsonify({'success': False, 'message': 'Email, password, and user name are required'}), 400

    # Load workbook, create if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(spreadsheet_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        # Create necessary columns
        worksheet = workbook.active
        worksheet['A1'] = 'Email'
        worksheet['B1'] = 'Password'
        worksheet['C1'] = 'User Name'

    worksheet = workbook.active
    for row in worksheet.iter_rows(min_row=2):
        if row[0].value == email:
            return jsonify({'success': False, 'message': 'User with this email already exists'}), 409

    # Hash the password
    hashed_password = hash_password(password)

    # Add the user to the spreadsheet
    row_index = worksheet.max_row + 1
    worksheet[f'A{row_index}'] = email
    worksheet[f'B{row_index}'] = hashed_password
    worksheet[f'C{row_index}'] = user_name

    workbook.save(spreadsheet_path)

    return jsonify({'success': True, 'message': 'User created successfully'})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
