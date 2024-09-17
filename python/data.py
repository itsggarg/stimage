from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
import datetime
import base64

app = Flask(__name__)
CORS(app)
app.config['JSON_AS_ASCII'] = False
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = True
app.config['JSON_SORT_KEYS'] = False

spreadsheet_path = 'DATA/stitched_data.xlsx'


@app.route('/save_data', methods=['POST'])
def save_data():
    # Now that we are using a JSON body, we must use request.json
    data = request.json

    course_code = data.get('course_code')
    classroom_number = data.get('classroom_number')
    face_count = data.get('face_count')
    image_path = data.get('image_path')
    user_name = data.get('user_name', 'Unknown User')
    stitched_image_base64 = data.get('stitched_image')
    device_ip = data.get('device_ip')
    device_mac = data.get('device_mac')
    device_name = data.get('device_name')

    # Load workbook, create if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(spreadsheet_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        # Create necessary columns
        worksheet = workbook.active
        worksheet['A1'] = 'Serial Number'
        worksheet['B1'] = 'Date'
        worksheet['C1'] = 'Time'
        worksheet['D1'] = 'Course Code'
        worksheet['E1'] = 'Classroom Number'
        worksheet['F1'] = 'User Name'
        worksheet['G1'] = 'Device Name'
        worksheet['H1'] = 'Device IP'
        worksheet['I1'] = 'Device MAC'
        worksheet['J1'] = 'Face Count'
        worksheet['K1'] = 'Image Path'
        worksheet['L1'] = 'Image Data'

    worksheet = workbook.active
    row_index = worksheet.max_row + 1  # Get the next available row

    # Get current date and time
    now = datetime.datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")

    # Save data to sheet
    worksheet[f'A{row_index}'] = row_index - 1  # Serial number
    worksheet[f'B{row_index}'] = date
    worksheet[f'C{row_index}'] = time
    worksheet[f'D{row_index}'] = course_code
    worksheet[f'E{row_index}'] = classroom_number
    worksheet[f'F{row_index}'] = user_name
    worksheet[f'G{row_index}'] = device_name
    worksheet[f'H{row_index}'] = device_ip
    worksheet[f'I{row_index}'] = device_mac
    worksheet[f'J{row_index}'] = face_count
    worksheet[f'K{row_index}'] = image_path
    worksheet[f'L{row_index}'] = stitched_image_base64

    # Save the workbook
    workbook.save(spreadsheet_path)

    return jsonify({'success': True, 'message': 'Data saved successfully'})


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5003)
